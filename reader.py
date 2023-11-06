import openpyxl

def printer(listToPrint):
    print(listToPrint)

def main(): 
    # Define variable to load the dataframe
    path = "./Book1.xlsx"

    dataframe = openpyxl.load_workbook(path)

    data = open("data.txt", "w")
    listing = open("moredata.txt", "w")
 
    # Define variable to read sheet
    dataframe1 = dataframe.active
    w = 1
    h = 43
    list1 = [[0 for x in range(w)] for y in range(h)]
    i = 0
    # Iterate the loop to read the cell values
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(0, dataframe1.max_column):
            data.write(str(col[row].value) + "\t\t\t")
            list1[i].append(str(col[row].value))
        data.write("\n")
        i+=1

    #print(str(list1))

    z = 0
    #create 2d list 
    while z < h:
        listing.write(str(list1[z]) + "\n")
        list1[z].pop(0)
        z += 1
    
    #printer(list1)
    data.close()
    listing.close()
    
   
    return list1

if __name__ == '__main__':
    main()


